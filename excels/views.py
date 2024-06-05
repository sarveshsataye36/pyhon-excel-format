import io
import os
from django.shortcuts import render,get_object_or_404,redirect
from django.conf import settings
import shutil
from django.http import JsonResponse, HttpResponse
import openpyxl
import re
import pandas as pd
from .models import Insurance,OfficeCode
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from datetime import date

def index(request):
    return render(request, 'excels/fileupload.html')


def upload_excel(request):
    if request.method == 'POST' and request.FILES['excel_file']:

        excel_file = request.FILES['excel_file']

        header_row_num = int(request.POST.get('header_row_num', 0))  # Get the starting row from the request, default is 0
        if(header_row_num > 0):
            header_row_num = header_row_num -1

        # Read the Excel file using pandas
        df = pd.read_excel(excel_file,skiprows=header_row_num)
        
        # Get column headers
        column_names = df.columns.tolist()

        # Construct the file path
        header_file_path = os.path.join(settings.BASE_DIR, 'data', 'header.xlsx')

        # Read the Excel file into a DataFrame
        header_excel = pd.read_excel(header_file_path)

        # Get column headers
        header_column_names = header_excel.columns.tolist()

        # Get Insurance column Name in dic format
        insurance_data = list(Insurance.objects.values('insurance_id', 'insurance_name'))

        return JsonResponse({'column_names': column_names, 'header_column': header_column_names, 'insurance_data':insurance_data})
    else:
        return JsonResponse({'error': 'No file was uploaded.'}, status=400)

def create_excel(request):
    if request.method == 'POST' and request.FILES['excel_file']:
        current_date = date.today().strftime('%d/%m/%Y')
        excel_file = request.FILES['excel_file']
        column_names = request.POST.getlist('selected_columns')
        header_row_num = int(request.POST.get('header_row_num', 0))
        insurance_id = request.POST.get('insurance_id')
        naration_date = request.POST.get('narationDate')
        naration_amount = request.POST.get('narationAmount')
        ntp_one = request.POST.get('ntp_one')
        cgst_checkbox = request.POST.get('cgst_checkbox')
        sgst_checkbox = request.POST.get('sgst_checkbox')
        header_list = request.POST.getlist('mapped_column')

        print(f"cgst_checkbox ==== {cgst_checkbox} || sgst_checkbox === {sgst_checkbox}")

        # Get the starting row from the request, default is 0
        if header_row_num > 0:
            header_row_num = header_row_num - 1

        df = pd.read_excel(excel_file, skiprows=header_row_num)
        filtered_df = df[column_names]

        # Remove specific characters from the filtered DataFrame
        filtered_df = filtered_df.applymap(lambda x: str(x).replace('`', '').replace(':', ''))

        # Load the formatted file from storage
        formatted_file_path = os.path.join(settings.BASE_DIR, 'data', 'header.xlsx')
        formatted_df = pd.read_excel(formatted_file_path)

        # Get single data from insurance table
        insurance = get_object_or_404(Insurance, insurance_id=insurance_id)
        insurance_name = insurance.insurance_name

        # Create a new DataFrame with the same columns as the formatted file
        merged_df = formatted_df.copy()

        # Copy the data from the filtered DataFrame to the corresponding columns in the merged DataFrame
        for col_name, header_name in zip(column_names, header_list):
            merged_df[header_name] = filtered_df[col_name]

        # Set default values
        merged_df['Debtor Name'] = insurance_name
        merged_df['Debtor Branch Ref'] = insurance_id
        merged_df['RepDate'] = current_date
        merged_df['NPT.1'] = ntp_one
        if cgst_checkbox != 'cgst_checkbox' and sgst_checkbox != 'sgst_checkbox':
            merged_df['Narration'] = f"BNG NEFT DT-{naration_date} rcvd towards brkg {naration_amount} from {insurance_name} with GST 18%"
        else:
            merged_df['Narration'] = f"BNG NEFT DT-{naration_date} rcvd towards brkg {naration_amount} from {insurance_name} without GST"

        # Create a new Excel file with pandas
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
            merged_df.to_excel(writer, index=False)

        # Save the Excel file to a temporary location
        temp_excel_file = 'temp_merged_data.xlsx'
        with pd.ExcelWriter(temp_excel_file, engine='openpyxl') as writer:
            merged_df.to_excel(writer, index=False)

        # Load the temporary file with openpyxl
        wb = load_workbook(temp_excel_file)
        ws = wb.active

        # Adjust column widths to fit the content
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter  # Get the column name (A, B, C, etc.)
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = max_length + 2  # Add some extra space
            ws.column_dimensions[column].width = adjusted_width

        # Save the modified workbook to BytesIO
        output = io.BytesIO()
        wb.save(output)

        # Create HTTP response with the new Excel file
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=merged_data.xlsx'

        # Remove the temporary file
        os.remove('temp_merged_data.xlsx')

        return response

    # If request method is GET or no file was uploaded, render the upload form
    return redirect('index')


def setting(request):
    return render(request, 'excels/setting.html')

def store_excel_data(request):
    if request.method == 'POST' and request.FILES['upload_excel_file']:
        excel_file = request.FILES['upload_excel_file']
        model_name = request.POST['model_name']

        uploaded_excel_to_db(excel_file,model_name)

        df = pd.read_excel(excel_file)
    
        return JsonResponse({'msg': 'Data upload sucess'}, status=200)
    else:
        return JsonResponse({'msg': 'No file was uploaded.'}, status=400)


def uploaded_excel_to_db(f, model_name):
    
    df = pd.read_excel(f)

    if model_name == 'insurance':

        # Truncate the table
        Insurance.objects.all().delete()
        # Insert new records
        for _, row in df.iterrows():
            Insurance.objects.create(
                insurance_id=row['Insurance_code'],
                insurance_name=row['Insurance_name']
            )

    elif  model_name == 'office_code':

        # Truncate the table
        OfficeCode.objects.all().delete()
        # Insert new records
        for _, row in df.iterrows():
            OfficeCode.objects.create(
                office_code_id=row['Office_code'],
                office_code_name=row['Office_branch']
            )




